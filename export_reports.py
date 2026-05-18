import json
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

# --- Data Loading and Filtering Logic ---
STANDARD_PORTS = {
    3389, 8080, 8443, 3306, 5432, 27017, 6379, 9000, 1521, 1433,
    5060, 5061, 8000, 8001, 8002, 8081, 8888, 9200, 5900, 10050, 10051
}

def is_standard_port(port):
    if port is None: return False
    return port <= 1024 or port in STANDARD_PORTS

def get_service_name(port):
    m = {22:'SSH', 80:'HTTP', 443:'HTTPS', 3389:'RDP', 3306:'MySQL', 5432:'PostgreSQL', 53:'DNS', 161:'SNMP'}
    return m.get(port, f"Port {port}")

print("Loading data...")
try:
    with open('notlike101_traffic_combined.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
        hits = data.get('hits', [])
except Exception as e:
    print(f"Error loading traffic data: {e}")
    hits = []

try:
    with open('host_map.json', 'r', encoding='utf-8') as f:
        host_map = json.load(f)
except:
    host_map = {}

def get_hostname(ip):
    return host_map.get(ip, {}).get('name', 'Unknown Host')

print("Processing traffic...")
host_centric = {}
service_centric = {}

for h in hits:
    dst_ip = h.get('dst_addr')
    src_ip = h.get('src_addr')
    dst_port = h.get('dst_port')
    proto = str(h.get('proto', 'TCP')).upper()

    if not dst_ip or not dst_ip.startswith('10.27.101.'): continue
    if src_ip == dst_ip: continue
    if not is_standard_port(dst_port): continue

    # Host-centric
    if dst_ip not in host_centric: host_centric[dst_ip] = {}
    if dst_port not in host_centric[dst_ip]: host_centric[dst_ip][dst_port] = {'proto': set(), 'sources': set()}
    host_centric[dst_ip][dst_port]['proto'].add(proto)
    host_centric[dst_ip][dst_port]['sources'].add(src_ip)

    # Service-centric
    if dst_port not in service_centric: service_centric[dst_port] = {'proto': set(), 'hosts': set()}
    service_centric[dst_port]['proto'].add(proto)
    service_centric[dst_port]['hosts'].add(dst_ip)

# --- Prepare DataFrames ---
print("Building DataFrames...")

# Build Host View Data
host_rows = []
for ip in sorted(host_centric.keys(), key=lambda x: int(x.split('.')[3])):
    hostname = get_hostname(ip)
    for port in sorted(host_centric[ip].keys()):
        stats = host_centric[ip][port]
        host_rows.append({
            "Target Host IP": ip,
            "Hostname": hostname,
            "Service Name": get_service_name(port),
            "Port Number": port,
            "Protocols": " / ".join(sorted(stats['proto'])),
            "Unique Sources Count": len(stats['sources']),
            "Decision (Allow/Deny)": "",
            "Remark": ""
        })

df_host = pd.DataFrame(host_rows)

# Build Service View Data
service_rows = []
for port in sorted(service_centric.keys()):
    stats = service_centric[port]
    for ip in sorted(stats['hosts'], key=lambda x: int(x.split('.')[3])):
        service_rows.append({
            "Port Number": port,
            "Service Name": get_service_name(port),
            "Protocols": " / ".join(sorted(stats['proto'])),
            "Target Host IP": ip,
            "Hostname": get_hostname(ip),
            "Decision (Valid/Invalid)": "",
            "Remark": ""
        })

df_service = pd.DataFrame(service_rows)

# --- Export CSV ---
print("Exporting CSVs...")
df_host.to_csv("Whitelist_By_Host.csv", index=False, encoding='utf-8-sig')
df_service.to_csv("Whitelist_By_Service.csv", index=False, encoding='utf-8-sig')

# --- Export Excel (.xlsx) with Styling ---
print("Exporting Excel...")
excel_file = "Whitelist_Reports.xlsx"
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    df_host.to_excel(writer, sheet_name="By Host", index=False)
    df_service.to_excel(writer, sheet_name="By Service", index=False)

    workbook = writer.book
    
    # Style configuration
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        # Style Headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # Adjust Column Widths
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            # Ensure max width of 50 to prevent huge columns
            worksheet.column_dimensions[col_letter].width = min(adjusted_width, 50)

# --- Export Print-Ready HTML (for PDF) ---
print("Exporting Print-Ready HTML...")
html_template = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Whitelist Report - Print Ready</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            color: #333;
            line-height: 1.4;
            margin: 20px;
        }}
        h1 {{ color: #1f4e78; text-align: center; font-size: 24px; }}
        h2 {{ color: #2e75b6; border-bottom: 2px solid #2e75b6; padding-bottom: 5px; margin-top: 30px; page-break-after: avoid; font-size: 18px; }}
        table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; font-size: 12px; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #1f4e78; color: white; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #f2f2f2; }}
        @media print {{
            @page {{ size: A4 landscape; margin: 1cm; }}
            body {{ margin: 0; }}
            .page-break {{ page-break-before: always; }}
            table {{ page-break-inside: auto; }}
            tr {{ page-break-inside: avoid; page-break-after: auto; }}
        }}
        .instruction {{ background: #fff3cd; color: #856404; padding: 10px; border: 1px solid #ffeeba; border-radius: 5px; margin-bottom: 20px; text-align: center; }}
        @media print {{ .instruction {{ display: none; }} }}
    </style>
</head>
<body>
    <div class="instruction">
        <strong>🖨️ วิธีเซฟเป็นไฟล์ PDF:</strong> เปิดไฟล์นี้บน Chrome หรือ Edge แล้วกด <code>Ctrl + P</code> หรือเมนู <strong>Print</strong> จากนั้นเลือก Destination เป็น <strong>"Save as PDF"</strong> (แนะนำให้เลือก Layout เป็น <strong>Landscape</strong>)
    </div>

    <h1>Zero-Trust Whitelist Cross-Validation Report</h1>

    <h2>View 1: Host-Centric Analysis (จัดกลุ่มตาม Server)</h2>
    {df_host.to_html(index=False, classes='table')}

    <div class="page-break"></div>

    <h2>View 2: Service-Centric Analysis (จัดกลุ่มตาม Port บริการ)</h2>
    {df_service.to_html(index=False, classes='table')}
</body>
</html>
"""

with open("Whitelist_Print_Ready.html", "w", encoding="utf-8-sig") as f:
    f.write(html_template)

print("Done! Generated .csv, .xlsx, and .html files.")
