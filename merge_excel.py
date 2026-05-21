import os
import shutil
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

def backup_file(file_path):
    backup_path = file_path + ".bak"
    if not os.path.exists(backup_path):
        print(f"Creating backup: {backup_path}")
        shutil.copy2(file_path, backup_path)
    else:
        print(f"Backup already exists: {backup_path}")

def style_sheet(file_path, sheet_name):
    print(f"Applying Teal styling & gridlines to sheet '{sheet_name}' in {os.path.basename(file_path)}...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    # Styling colors
    teal_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Enable gridlines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Style the header row (row 1)
    for cell in ws[1]:
        cell.fill = teal_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
    # Set standard row height for header
    ws.row_dimensions[1].height = 26
    
    # Auto-fit column widths (capping scan at 100 rows to ensure instantaneous completion)
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for i, cell in enumerate(col):
            if i >= 100:
                break
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)
        
    wb.save(file_path)
    print("Styling completed successfully!")

def merge_file_b():
    file_b = r"e:\vibe_code\get_project_craft_acl\B-MUIC Sevice Port - Categorize Reference.xlsx"
    print("\n==================================================")
    print(f"Processing File B: {os.path.basename(file_b)}")
    print("==================================================")
    
    if not os.path.exists(file_b):
        print(f"Error: File B not found at {file_b}")
        return
        
    backup_file(file_b)
    
    print("Reading sheets...")
    df_std = pd.read_excel(file_b, sheet_name="1. Standard Ports (Allowed)")
    df_drop = pd.read_excel(file_b, sheet_name="2. Dropped Ports (Review)")
    
    print(f"Loaded {len(df_std)} standard rows and {len(df_drop)} dropped rows.")
    
    # Align Standard ports schema
    df1 = pd.DataFrame()
    df1["Category"] = ["Allowed (Standard)"] * len(df_std)
    df1["Target Host IP"] = ["* (All Hosts)"] * len(df_std)
    df1["Hostname"] = ["Approved Enterprise Ports"] * len(df_std)
    df1["Port Number"] = df_std["Port Number"]
    df1["Protocols"] = df_std["Protocols"]
    df1["Unique Sources Count"] = df_std["Total Unique Sources"]
    df1["Details / Reason"] = df_std["Service Type"].astype(str) + " - " + df_std["Reason"].astype(str)
    
    # Align Dropped ports schema
    df2 = pd.DataFrame()
    df2["Category"] = ["Dropped (For Review)"] * len(df_drop)
    df2["Target Host IP"] = df_drop["Target Host IP"]
    df2["Hostname"] = df_drop["Hostname"]
    df2["Port Number"] = df_drop["Dropped Port Number"]
    df2["Protocols"] = df_drop["Protocols"]
    df2["Unique Sources Count"] = df_drop["Unique Sources Count"]
    df2["Details / Reason"] = df_drop["Potential Reason"].astype(str) + " (" + df_drop["Review Required"].astype(str) + ")"
    
    # Merge
    print("Merging DataFrames...")
    df_merged = pd.concat([df1, df2], ignore_index=True)
    sheet_name = "Consolidated Port Reference"
    
    print(f"Saving merged data ({len(df_merged)} rows) to {file_b}...")
    try:
        with pd.ExcelWriter(file_b, engine='openpyxl') as writer:
            df_merged.to_excel(writer, sheet_name=sheet_name, index=False)
        print("Write successful!")
        style_sheet(file_b, sheet_name)
    except PermissionError:
        print(f"\n❌ Error: Permission denied writing to {file_b}.")
        print("Please close the Excel file if it is open in Microsoft Excel and run the script again.")
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")

def merge_file_c():
    file_c = r"e:\vibe_code\get_project_craft_acl\C-Port_Triage_For_Meeting.xlsx"
    print("\n==================================================")
    print(f"Processing File C: {os.path.basename(file_c)}")
    print("==================================================")
    
    if not os.path.exists(file_c):
        print(f"Error: File C not found at {file_c}")
        return
        
    backup_file(file_c)
    
    print("Reading sheets...")
    df_action = pd.read_excel(file_c, sheet_name="1. Action Required (Review)")
    df_noise = pd.read_excel(file_c, sheet_name="2. Ignored Noise (Ephemeral)")
    
    print(f"Loaded {len(df_action)} action rows and {len(df_noise)} noise rows.")
    
    rows_c = []
    
    # Helper to check if a dataframe has actual data rows (not just a placeholder "No custom ports found" row)
    def is_valid_df(df):
        if df.empty:
            return False
        if "Message" in df.columns:
            return False
        return "Target Host IP" in df.columns

    if is_valid_df(df_action):
        df_action_clean = df_action.copy()
        df_action_clean.insert(0, "Triage Category", "Action Required (Review)")
        rows_c.append(df_action_clean)
        print("Added Action Required data.")
    else:
        print("Skipping empty/placeholder Action Required sheet.")
        
    if is_valid_df(df_noise):
        df_noise_clean = df_noise.copy()
        df_noise_clean.insert(0, "Triage Category", "Ignored Noise (Ephemeral)")
        rows_c.append(df_noise_clean)
        print("Added Ignored Noise data.")
    else:
        print("Skipping empty/placeholder Ignored Noise sheet.")
        
    if rows_c:
        df_merged = pd.concat(rows_c, ignore_index=True)
    else:
        df_merged = pd.DataFrame([{"Message": "No data found."}])
        
    sheet_name = "Consolidated Triage Reference"
    
    print(f"Saving merged data ({len(df_merged)} rows) to {file_c}...")
    try:
        with pd.ExcelWriter(file_c, engine='openpyxl') as writer:
            df_merged.to_excel(writer, sheet_name=sheet_name, index=False)
        print("Write successful!")
        style_sheet(file_c, sheet_name)
    except PermissionError:
        print(f"\n❌ Error: Permission denied writing to {file_c}.")
        print("Please close the Excel file if it is open in Microsoft Excel and run the script again.")
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")

if __name__ == "__main__":
    merge_file_b()
    merge_file_c()
    print("\n🎉 Consolidation completed successfully!")
