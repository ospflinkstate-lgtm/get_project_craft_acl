import re
import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Port mapping to exact rule orders and service names
PORT_RULE_MAP = {
    22: ("[05]", "SSH (22)"),
    25: ("[06]", "SMTP (25)"),
    53: ("[07]", "DNS (53)"),
    80: ("[08]", "HTTP (80)"),
    88: ("[09]", "Kerberos (88)"),
    111: ("[10]", "RPCBind (111)"),
    135: ("[11]", "MS-RPC (135)"),
    139: ("[12]", "NetBIOS (139)"),
    389: ("[13]", "LDAP (389)"),
    443: ("[14]", "HTTPS (443)"),
    445: ("[15]", "SMB (445)"),
    853: ("[16]", "DNS-over-TLS (853)"),
    1024: ("[17]", "MS-RPC-Dynamic (1024)"),
    1433: ("[18]", "MSSQL (1433)"),
    1521: ("[19]", "Oracle DB (1521)"),
    3389: ("[20]", "RDP (3389)"),
    8000: ("[21]", "HTTP-Alt (8000)"),
    8002: ("[22]", "Custom (8002)"),
    8080: ("[23]", "HTTP-Proxy (8080)"),
    8081: ("[24]", "Custom (8081)"),
    8443: ("[25]", "HTTPS-Alt (8443)"),
    8888: ("[26]", "Custom (8888)"),
    9000: ("[27]", "Custom (9000)"),
    9200: ("[28]", "Elasticsearch (9200)"),
    10051: ("[29]", "Zabbix Agent (10051)")
}

def parse_mikrotik_rsc(filepath):
    hosts = {}  # ip -> {ip, hostname, lists: [], ports: [], status, phase, rules: []}
    
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        
    # Unify wrapped lines in RouterOS export (lines ending with \)
    unified_lines = []
    current_line = ""
    for line in lines:
        line_str = line.strip('\r\n')
        if line_str.endswith('\\'):
            current_line += line_str[:-1]
        else:
            current_line += line_str
            unified_lines.append(current_line)
            current_line = ""
    if current_line:
        unified_lines.append(current_line)
        
    current_section = ""
    
    for line in unified_lines:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
            
        # Track current configuration section
        if line.startswith('/'):
            current_section = line
            continue
            
        if "ip firewall address-list" in current_section and line.startswith("add"):
            # Robust parameter extraction (allowing arbitrary whitespace around equals and quotes)
            ip_match = re.search(r'address=\s*([0-9\./\w]+)', line)
            list_match = re.search(r'list=\s*"([^"]+)"', line)
            if not list_match:
                list_match = re.search(r'list=\s*([^ ]+)', line)
            
            # Quoted comments first, then unquoted
            comment_match = re.search(r'comment=\s*"([^"]+)"', line)
            if not comment_match:
                comment_match = re.search(r'comment=\s*([^ ]+)', line)
                
            if ip_match and list_match:
                ip_full = ip_match.group(1)
                ip = ip_full.split('/')[0] # Remove subnet mask
                
                # Verify valid IPv4
                if not re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', ip):
                    continue
                    
                # Only track hosts in our main VLAN 21 subnet (10.27.101.X)
                if not ip.startswith("10.27.101."):
                    continue
                    
                lst = list_match.group(1).strip('"')
                comment = "Unknown Host"
                if comment_match:
                    comment = comment_match.group(1).strip('"')
                    
                if ip not in hosts:
                    hosts[ip] = {
                        "ip": ip,
                        "hostname": comment,
                        "lists": [],
                        "ports": [],
                        "phase": "Phase 2 (Production)",
                        "status": "Bypassed (Safe Mode)",  # Default
                        "policies": "",
                        "rules": []
                    }
                
                if comment != "Unknown Host" and (hosts[ip]["hostname"] == "Unknown Host" or not hosts[ip]["hostname"]):
                    hosts[ip]["hostname"] = comment
                    
                if lst not in hosts[ip]["lists"]:
                    hosts[ip]["lists"].append(lst)
                    
                # Check for allow-port-X-servers list
                port_match = re.match(r'allow-port-(\d+)-servers', lst)
                if port_match:
                    port = int(port_match.group(1))
                    if port not in hosts[ip]["ports"]:
                        hosts[ip]["ports"].append(port)

    # Step 2: Manually customize the PoC Target Hosts (Phase 1)
    if "10.27.101.41" in hosts:
        hosts["10.27.101.41"]["hostname"] = "Safeline WAF"
        hosts["10.27.101.41"]["phase"] = "Phase 1 (PoC Target)"
        hosts["10.27.101.41"]["status"] = "Enforced (Active)"
        hosts["10.27.101.41"]["policies"] = "Zero-Trust PoC, Allowed: HTTP (80), HTTPS (443), SSH (22) & Web Admin (9443) from Admin VLAN only, ICMP (Ping). Blocked: All other ports."
        hosts["10.27.101.41"]["rules"] = "[01], [02], [02-A], [03], [04], [99]"

    if "10.27.101.165" in hosts:
        hosts["10.27.101.165"]["hostname"] = "Backend Apache Server"
        hosts["10.27.101.165"]["phase"] = "Phase 1 (PoC Target)"
        hosts["10.27.101.165"]["status"] = "Enforced (Active)"
        hosts["10.27.101.165"]["policies"] = "Zero-Trust PoC, Allowed: Proxy Port 8881 (from WAF only), SSH (22) from Admin VLAN only, ICMP (Ping). Blocked: All other ports."
        hosts["10.27.101.165"]["rules"] = "[01], [02], [02-A], [02-B], [04], [99]"

    # Step 3: Populate details for Phase 2 Production Hosts
    for ip, h in hosts.items():
        if ip in ["10.27.101.41", "10.27.101.165"]:
            continue
            
        h["phase"] = "Phase 2 (Production)"
        h["status"] = "Bypassed (Safe Mode)"
        
        srv_list = []
        rule_list = ["[01]", "[02]"]
        
        sorted_ports = sorted(h["ports"])
        for p in sorted_ports:
            if p in PORT_RULE_MAP:
                rule, name = PORT_RULE_MAP[p]
                srv_list.append(name)
                if rule not in rule_list:
                    rule_list.append(rule)
            else:
                srv_list.append(f"Port {p}")
                
        # Sort rule list numerically to ensure clean formatting
        def get_rule_num(r):
            try:
                return float(r.strip("[]").split("-")[0])
            except ValueError:
                return 999.0
        rule_list = sorted(list(set(rule_list)), key=get_rule_num)
        
        h["policies"] = f"Bypassed from Zero-Trust Block. Allowed Services: {', '.join(srv_list) if srv_list else 'Standard L2 Bridge Traffic'}"
        h["rules"] = ", ".join(rule_list)

    # Sort hosts: first by phase (Phase 1 first), then by IP host octet numerically
    return sorted(hosts.values(), key=lambda x: (x["phase"], int(x["ip"].split(".")[3])))

def parse_filter_rules(filepath):
    rules = []
    
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        
    unified_lines = []
    current_line = ""
    for line in lines:
        line_str = line.strip('\r\n')
        if line_str.endswith('\\'):
            current_line += line_str[:-1]
        else:
            current_line += line_str
            unified_lines.append(current_line)
            current_line = ""
    if current_line:
        unified_lines.append(current_line)
        
    current_section = ""
    for line in unified_lines:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
            
        if line.startswith('/'):
            current_section = line
            continue
            
        if "ip firewall filter" in current_section and line.startswith("add"):
            chain_match = re.search(r'chain=\s*([^ ]+)', line)
            action_match = re.search(r'action=\s*([^ ]+)', line)
            proto_match = re.search(r'protocol=\s*([^ ]+)', line)
            port_match = re.search(r'dst-port=\s*([^ ]+)', line)
            
            comment_match = re.search(r'comment=\s*"([^"]+)"', line)
            if not comment_match:
                comment_match = re.search(r'comment=\s*([^ ]+)', line)
                
            src_list_match = re.search(r'src-address-list=\s*"([^"]+)"', line)
            if not src_list_match:
                src_list_match = re.search(r'src-address-list=\s*([^ ]+)', line)
                
            dst_list_match = re.search(r'dst-address-list=\s*"([^"]+)"', line)
            if not dst_list_match:
                dst_list_match = re.search(r'dst-address-list=\s*([^ ]+)', line)
                
            src_addr_match = re.search(r'src-address=\s*([^ ]+)', line)
            dst_addr_match = re.search(r'dst-address=\s*([^ ]+)', line)
            
            comment = ""
            if comment_match:
                comment = comment_match.group(1).strip('"')
                
            # Extract rule order
            order = "-"
            desc = comment
            order_match = re.match(r'^\[([a-zA-Z0-9\-]+)\]\s*(.*)$', comment)
            if order_match:
                order = f"[{order_match.group(1)}]"
                desc = order_match.group(2)
            else:
                # If the comment doesn't have [XX], it's not our structured rule
                continue
                
            chain = chain_match.group(1).strip('"') if chain_match else "-"
            action = action_match.group(1).strip('"') if action_match else "-"
            protocol = proto_match.group(1).strip('"') if proto_match else "-"
            dst_port = port_match.group(1).strip('"') if port_match else "-"
            
            src_list = src_list_match.group(1).strip('"') if src_list_match else ""
            dst_list = dst_list_match.group(1).strip('"') if dst_list_match else ""
            src_addr = src_addr_match.group(1).strip('"') if src_addr_match else ""
            dst_addr = dst_addr_match.group(1).strip('"') if dst_addr_match else ""
            
            addr_list_str = "-"
            if src_list and dst_list:
                addr_list_str = f"src: {src_list} -> dst: {dst_list}"
            elif src_list:
                addr_list_str = f"src: {src_list}"
            elif dst_list:
                addr_list_str = f"dst: {dst_list}"
            elif src_addr and dst_addr:
                addr_list_str = f"src: {src_addr} -> dst: {dst_addr}"
            elif src_addr:
                addr_list_str = f"src: {src_addr}"
            elif dst_addr:
                addr_list_str = f"dst: {dst_addr}"
                
            # Determine target subnet
            target = "All Subnets"
            if "poc-targets" in line or "10.27.101.41" in line or "10.27.101.165" in line:
                target = "PoC Targets"
            elif "allow-port" in line:
                target = "Production Servers"
                
            rules.append({
                "Rule Order": order,
                "Chain": chain,
                "Action": action,
                "Protocol": protocol,
                "Dst. Port": dst_port,
                "Source / Dst Address List": addr_list_str,
                "Rule Description": desc,
                "Applied Target": target
            })
            
    # Sort rules
    def get_rule_key(r):
        o = r["Rule Order"].strip("[]")
        parts = o.split('-')
        try:
            val = float(parts[0])
        except ValueError:
            val = 999.0
        if len(parts) > 1:
            let = parts[1]
            let_val = ord(let.lower()) - 96
            val += let_val * 0.01
        return val
        
    return sorted(rules, key=get_rule_key)

def generate_excel(rsc_file="Mikrotik-TP2.rsc"):
    excel_file = "MikroTik_PoC_Enforcement_Tracker.xlsx"
    
    if not os.path.exists(rsc_file):
        print(f"Error: File '{rsc_file}' not found in the current directory!")
        print("Please make sure the file path is correct.")
        return
        
    print(f"Parsing configuration from {rsc_file}...")
    hosts_data = parse_mikrotik_rsc(rsc_file)
    rules_data = parse_filter_rules(rsc_file)
    
    # Sheet 1 DataFrame: Host Tracker
    tracker_rows = []
    for idx, h in enumerate(hosts_data, 1):
        tracker_rows.append({
            "No.": idx,
            "IP Address": h["ip"],
            "Hostname / Server Description": h["hostname"],
            "PoC Phase": h["phase"],
            "Enforcement Status": h["status"],
            "Applied Firewall Policies": h["policies"],
            "Associated Firewall Rules": h["rules"]
        })
    df_tracker = pd.DataFrame(tracker_rows)
    
    # Sheet 2 DataFrame: Filter Rules Reference
    df_rules = pd.DataFrame(rules_data)
    
    print(f"Writing data to Excel sheet {excel_file}...")
    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df_tracker.to_excel(writer, sheet_name="PoC Host Tracker", index=False)
            df_rules.to_excel(writer, sheet_name="Filter Rules Reference", index=False)
            
            # Style Sheets
            style_excel_sheets(writer)
    except PermissionError:
        print(f"\n[ERROR] Cannot write to '{excel_file}'.")
        print("Please close the Excel file in Microsoft Excel or any other program and try again!")
        return
        
    print(f"Excel generation completed successfully! Total hosts tracked: {len(hosts_data)}, Total rules listed: {len(rules_data)}")

def style_excel_sheets(writer):
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid") # Teal header
    active_poc_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Light Green for Enforced
    bypass_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Light Blue for Bypassed
    
    header_font = Font(name="Segoe UI", size=11, color="FFFFFF", bold=True)
    body_font = Font(name="Segoe UI", size=10, color="333333")
    bold_body_font = Font(name="Segoe UI", size=10, bold=True, color="333333")
    
    status_enforced_font = Font(name="Segoe UI", size=10, color="065F46", bold=True)
    status_bypass_font = Font(name="Segoe UI", size=10, color="1E40AF", bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Sheet 1 Formatting: PoC Host Tracker
    ws_tracker = writer.sheets["PoC Host Tracker"]
    ws_tracker.views.sheetView[0].showGridLines = True
    
    for cell in ws_tracker[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws_tracker.row_dimensions[1].height = 28
    
    for row in range(2, ws_tracker.max_row + 1):
        ws_tracker.row_dimensions[row].height = 20
        status_val = ws_tracker.cell(row=row, column=5).value
        
        for col in range(1, ws_tracker.max_column + 1):
            cell = ws_tracker.cell(row=row, column=col)
            cell.font = body_font
            cell.border = thin_border
            
            if col in [1, 2, 4]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
        cell_status = ws_tracker.cell(row=row, column=5)
        cell_ip = ws_tracker.cell(row=row, column=2)
        cell_host = ws_tracker.cell(row=row, column=3)
        
        if status_val == "Enforced (Active)":
            cell_status.fill = active_poc_fill
            cell_status.font = status_enforced_font
            cell_ip.font = bold_body_font
            cell_host.font = bold_body_font
        elif status_val == "Bypassed (Safe Mode)":
            cell_status.fill = bypass_fill
            cell_status.font = status_bypass_font
            
    autofit_columns(ws_tracker)
    
    # Sheet 2 Formatting: Filter Rules Reference
    ws_rules = writer.sheets["Filter Rules Reference"]
    ws_rules.views.sheetView[0].showGridLines = True
    
    for cell in ws_rules[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
    ws_rules.row_dimensions[1].height = 28
    
    for row in range(2, ws_rules.max_row + 1):
        ws_rules.row_dimensions[row].height = 24
        rule_order = ws_rules.cell(row=row, column=1).value
        
        for col in range(1, ws_rules.max_column + 1):
            cell = ws_rules.cell(row=row, column=col)
            cell.font = body_font
            cell.border = thin_border
            
            if col in [1, 2, 3, 4, 5]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
        cell_rule = ws_rules.cell(row=row, column=1)
        if rule_order in ["[02-B]", "[03]", "[04]"]:
            cell_rule.font = bold_body_font
        elif rule_order == "[99]":
            cell_rule.font = Font(name="Segoe UI", size=10, bold=True, color="990000")
            cell_rule.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            
    autofit_columns(ws_rules)

def autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)
            except:
                pass
        adjusted_width = min(max((max_length + 3), 10), 60)
        ws.column_dimensions[col_letter].width = adjusted_width

if __name__ == "__main__":
    target_file = "Mikrotik-TP3.rsc"
    if len(sys.argv) > 1:
        target_file = sys.argv[1]
        
    generate_excel(target_file)
