import json
import pandas as pd
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

STANDARD_PORTS = {
    3389, 8080, 8443, 3306, 5432, 27017, 6379, 9000, 1521, 1433,
    5060, 5061, 8000, 8001, 8002, 8081, 8888, 9200, 5900, 10050, 10051
}

def is_standard_port(port):
    if port is None: return False
    return port <= 1024 or port in STANDARD_PORTS

def is_cluster_node(hostname):
    h = hostname.lower()
    return any(x in h for x in ['swarm', 'k8s', 'k3s', 'node'])

print("Loading data...")
try:
    with open('notlike101_traffic_combined.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
        hits = data.get('hits', [])
except Exception as e:
    print(f"Error loading traffic data: {e}")
    hits = []

try:
    with open('host_map.json', 'r', encoding='utf-8') as f:
        host_map = json.load(f)
except:
    host_map = {}

def get_hostname(ip):
    return host_map.get(ip, {}).get('name', 'Unknown Host')

print("Applying Triage Logic to Dropped Ports...")

# dropped_traffic: dict of (dst_ip, dst_port) -> { proto: set, unique_sources: set }
dropped_traffic = defaultdict(lambda: {'proto': set(), 'unique_sources': set()})

for h in hits:
    dst_ip = h.get('dst_addr')
    src_ip = h.get('src_addr')
    dst_port = h.get('dst_port')
    proto = str(h.get('proto', 'TCP')).upper()

    if not dst_ip or not dst_ip.startswith('10.27.101.'): continue
    if src_ip == dst_ip: continue
    if dst_port is None: continue

    if not is_standard_port(dst_port):
        key = (dst_ip, dst_port)
        dropped_traffic[key]['proto'].add(proto)
        dropped_traffic[key]['unique_sources'].add(src_ip)

action_required = []
noise_ignored = []

for (dst_ip, dst_port) in sorted(dropped_traffic.keys(), key=lambda x: (int(x[0].split('.')[3]), x[1])):
    stats = dropped_traffic[(dst_ip, dst_port)]
    sources_count = len(stats['unique_sources'])
    hostname = get_hostname(dst_ip)
    
    row_data = {
        "Target Host IP": dst_ip,
        "Hostname": hostname,
        "Target Port": dst_port,
        "Protocols": ", ".join(sorted(stats['proto'])),
        "Unique Sources Count": sources_count,
    }

    if is_cluster_node(hostname):
        row_data["AI Recommendation"] = "Cluster Node - Recommend allowing Range (e.g., 10000-40000)"
        row_data["Confidence"] = "High (Cluster Architecture)"
        action_required.append(row_data)
    elif sources_count >= 5:
        row_data["AI Recommendation"] = "Likely Custom Service - Needs Manual Review"
        row_data["Confidence"] = "Medium (High Hit Rate)"
        action_required.append(row_data)
    else:
        row_data["AI Recommendation"] = "Likely Ephemeral (Return Traffic) / Scan - Safe to Ignore"
        row_data["Confidence"] = "High (Low Hit Rate)"
        noise_ignored.append(row_data)

print("Preparing DataFrames...")
df_action = pd.DataFrame(action_required)
df_noise = pd.DataFrame(noise_ignored)

excel_file = "Port_Triage_For_Meeting.xlsx"
print(f"Exporting to {excel_file}...")

with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    if not df_action.empty:
        df_action.to_excel(writer, sheet_name="1. Action Required (Review)", index=False)
    else:
        pd.DataFrame([{"Message": "No high-probability custom ports found."}]).to_excel(writer, sheet_name="1. Action Required (Review)", index=False)
        
    if not df_noise.empty:
        df_noise.to_excel(writer, sheet_name="2. Ignored Noise (Ephemeral)", index=False)
    else:
        pd.DataFrame([{"Message": "No noise ports found."}]).to_excel(writer, sheet_name="2. Ignored Noise (Ephemeral)", index=False)

    workbook = writer.book
    
    action_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid") # Amber
    noise_fill = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid") # Slate
    header_font = Font(color="FFFFFF", bold=True)
    
    if not df_action.empty:
        ws1 = writer.sheets["1. Action Required (Review)"]
        for cell in ws1[1]:
            cell.fill = action_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ws1.columns:
            col_letter = get_column_letter(col[0].column)
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws1.column_dimensions[col_letter].width = min(max_length + 2, 50)

    if not df_noise.empty:
        ws2 = writer.sheets["2. Ignored Noise (Ephemeral)"]
        for cell in ws2[1]:
            cell.fill = noise_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ws2.columns:
            col_letter = get_column_letter(col[0].column)
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws2.column_dimensions[col_letter].width = min(max_length + 2, 50)

print("Done!")
