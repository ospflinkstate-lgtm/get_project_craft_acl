"""
Update A-Draft - Mikrotik Deployment.xlsx from current Mikrotik-TP2.rsc config.
Parses the RSC file (handling multi-line continuations) and regenerates
Sheet 1: Address Lists  and  Sheet 2: Filter Rules
with premium Teal (#0F766E) styling.
"""
import re
import sys
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

RSC_FILE = r"e:\vibe_code\get_project_craft_acl\Mikrotik-TP2.rsc"
OUTPUT_FILE = r"e:\vibe_code\get_project_craft_acl\A-Draft - Mikrotik Deployment.xlsx"

# ── Step 1: Read RSC and join continuation lines ──
def read_rsc(path):
    with open(path, 'r', encoding='utf-8') as f:
        raw = f.read()
    # Join lines ending with backslash-continuation
    raw = raw.replace('\r\n', '\n')
    raw = re.sub(r'\\\s*\n\s*', ' ', raw)
    return raw.split('\n')

# ── Step 2: Parse address-list entries ──
def parse_address_lists(lines):
    rows = []
    counter = 0
    in_section = False
    for line in lines:
        if line.startswith('/ip firewall address-list'):
            in_section = True
            continue
        if line.startswith('/') and in_section:
            break
        if in_section and line.strip().startswith('add '):
            counter += 1
            # Extract fields
            addr_m = re.search(r'address=(\S+)', line)
            comment_m = re.search(r'comment=("([^"]*)"|([\S]+))', line)
            list_m = re.search(r'list=(\S+)', line)

            ip = addr_m.group(1) if addr_m else ''
            if comment_m:
                hostname = comment_m.group(2) if comment_m.group(2) else comment_m.group(3)
            else:
                hostname = ''
            list_name = list_m.group(1) if list_m else ''

            # Derive target service from list name
            port_m = re.search(r'allow-port-(\d+)', list_name)
            if port_m:
                service = f"Port {port_m.group(1)}"
            elif list_name == 'poc-targets':
                service = 'PoC Target (Drop Zone)'
            elif list_name == 'waf-frontend':
                service = 'WAF Frontend (80/443)'
            elif list_name == 'admin-management':
                service = 'Admin Management (SSH)'
            else:
                service = list_name

            # Build the MikroTik command
            cmd = line.strip()

            rows.append({
                'No.': counter,
                'Address List Name': list_name,
                'IP Address': ip,
                'Hostname (Comment)': hostname,
                'Target Service': service,
                'MikroTik Command (Copy & Paste)': cmd
            })
    return rows

# ── Step 3: Parse filter rules ──
def parse_filter_rules(lines):
    rows = []
    in_section = False
    for line in lines:
        if line.startswith('/ip firewall filter'):
            in_section = True
            continue
        if line.startswith('/') and in_section:
            break
        if in_section and line.strip().startswith('add '):
            # Extract fields
            action_m = re.search(r'action=(\S+)', line)
            chain_m = re.search(r'chain=(\S+)', line)
            comment_m = re.search(r'comment=("([^"]*)"|([\S]+))', line)
            proto_m = re.search(r'protocol=(\S+)', line)
            port_m = re.search(r'dst-port=(\S+)', line)
            dst_list_m = re.search(r'dst-address-list=(\S+)', line)
            dst_addr_m = re.search(r'dst-address=(\S+)', line)
            conn_m = re.search(r'connection-state=(\S+)', line)

            action = action_m.group(1) if action_m else ''
            chain = chain_m.group(1) if chain_m else ''
            if comment_m:
                comment = comment_m.group(2) if comment_m.group(2) else comment_m.group(3)
            else:
                comment = ''
            protocol = proto_m.group(1).upper() if proto_m else ''
            port = port_m.group(1) if port_m else ''
            
            # Target: prefer address-list, then direct address, then connection-state
            if dst_list_m:
                target = dst_list_m.group(1)
            elif dst_addr_m:
                target = dst_addr_m.group(1)
            elif conn_m:
                target = f"connection-state={conn_m.group(1)}"
            else:
                target = '(all traffic)'

            # Extract rule order from comment like [01], [02-A], [99]
            order_m = re.search(r'\[([^\]]+)\]', comment)
            rule_order = order_m.group(1) if order_m else ''

            cmd = line.strip()

            rows.append({
                'Rule Order': rule_order,
                'Chain': chain,
                'Action': action.upper(),
                'Protocol': protocol if protocol else '-',
                'Port': port if port else '-',
                'Target Address-List': target,
                'Comment / Description': comment,
                'MikroTik Command (Copy & Paste)': cmd
            })
    return rows

# ── Step 4: Style a worksheet with premium Teal theme ──
def style_sheet(ws, header_color="0F766E"):
    teal_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, color="FFFFFF", bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # Style header row
    for cell in ws[1]:
        cell.fill = teal_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    # Style data rows (light font + borders)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=False)

    # Auto-fit column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for i, cell in enumerate(col):
            if i >= 200:
                break
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 80)

    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True

# ── Main ──
def main():
    target_file = RSC_FILE
    if len(sys.argv) > 1:
        target_file = sys.argv[1]
    print(f"Reading {target_file}...")
    lines = read_rsc(target_file)
    
    print("Parsing address lists...")
    addr_rows = parse_address_lists(lines)
    print(f"  Found {len(addr_rows)} address-list entries.")
    
    print("Parsing filter rules...")
    rule_rows = parse_filter_rules(lines)
    print(f"  Found {len(rule_rows)} filter rules.")
    
    df_addr = pd.DataFrame(addr_rows)
    df_rules = pd.DataFrame(rule_rows)
    
    print(f"Writing to {OUTPUT_FILE}...")
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df_addr.to_excel(writer, sheet_name="1. Address Lists", index=False)
        df_rules.to_excel(writer, sheet_name="2. Filter Rules", index=False)
    
    # Apply styling
    print("Applying premium Teal styling...")
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    for sheet_name in wb.sheetnames:
        style_sheet(wb[sheet_name])
    wb.save(OUTPUT_FILE)
    
    print(f"\nDone! Updated '{OUTPUT_FILE}'")
    print(f"   Sheet '1. Address Lists':  {len(addr_rows)} rows")
    print(f"   Sheet '2. Filter Rules':   {len(rule_rows)} rules")

if __name__ == "__main__":
    main()
