import json
import pandas as pd
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

STANDARD_PORTS = {
    3389, 8080, 8443, 3306, 5432, 27017, 6379, 9000, 1521, 1433,
    5060, 5061, 8000, 8001, 8002, 8081, 8888, 9200, 5900, 10050, 10051
}

def is_standard_port(port):
    if port is None: return False
    return port <= 1024 or port in STANDARD_PORTS

def get_service_name(port):
    m = {22:'SSH', 80:'HTTP', 443:'HTTPS', 3389:'RDP', 3306:'MySQL', 5432:'PostgreSQL', 53:'DNS', 161:'SNMP'}
    return m.get(port, f"Port_{port}")

print("Loading data...")
try:
    with open('notlike101_traffic_combined.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
        hits = data.get('hits', [])
except Exception as e:
    print(f"Error loading traffic data: {e}")
    hits = []

try:
    with open('host_map.json', 'r', encoding='utf-8') as f:
        host_map = json.load(f)
except:
    host_map = {}

def get_hostname(ip):
    return host_map.get(ip, {}).get('name', 'Unknown_Host').replace(' ', '_').replace('(', '').replace(')', '')

print("Processing traffic for Excel Deployment Plan...")
service_centric = defaultdict(lambda: {'proto': set(), 'hosts': set()})

for h in hits:
    dst_ip = h.get('dst_addr')
    src_ip = h.get('src_addr')
    dst_port = h.get('dst_port')
    proto = str(h.get('proto', 'TCP')).lower()

    if not dst_ip or not dst_ip.startswith('10.27.101.'): continue
    if src_ip == dst_ip: continue
    if not is_standard_port(dst_port): continue

    service_centric[dst_port]['proto'].add(proto)
    service_centric[dst_port]['hosts'].add(dst_ip)

# Prepare data for Address Lists Sheet
address_list_data = []
sorted_ports = sorted(service_centric.keys())

counter = 1
for port in sorted_ports:
    srv_name = get_service_name(port)
    hosts = sorted(list(service_centric[port]['hosts']), key=lambda x: int(x.split('.')[3]))
    list_name = f"allow-port-{port}-servers"
    
    for ip in hosts:
        hostname = get_hostname(ip)
        command = f"add list=\"{list_name}\" address={ip} comment=\"{hostname}\""
        address_list_data.append({
            "No.": counter,
            "Address List Name": list_name,
            "IP Address": ip,
            "Hostname (Comment)": hostname,
            "Target Service": srv_name,
            "MikroTik Command (Copy & Paste)": command
        })
        counter += 1

# Prepare data for Filter Rules Sheet
filter_rules_data = []
filter_rules_data.append({
    "Rule Order": "01",
    "Chain": "forward",
    "Action": "accept",
    "Protocol": "-",
    "Port": "-",
    "Target Address-List": "-",
    "Comment / Description": "[01] ACCEPT ESTABLISHED/RELATED",
    "MikroTik Command (Copy & Paste)": "add chain=forward action=accept connection-state=established,related comment=\"[01] ACCEPT ESTABLISHED/RELATED\""
})
filter_rules_data.append({
    "Rule Order": "02",
    "Chain": "forward",
    "Action": "drop",
    "Protocol": "-",
    "Port": "-",
    "Target Address-List": "-",
    "Comment / Description": "[02] DROP INVALID",
    "MikroTik Command (Copy & Paste)": "add chain=forward action=drop connection-state=invalid comment=\"[02] DROP INVALID\""
})

rule_idx = 3
for port in sorted_ports:
    srv_name = get_service_name(port)
    protos = list(service_centric[port]['proto'])
    list_name = f"allow-port-{port}-servers"
    
    for p in protos:
        comment = f"[{rule_idx:02d}] ALLOW {srv_name.upper()} ({p.upper()})"
        command = f"add chain=forward action=accept protocol={p} dst-port={port} dst-address-list=\"{list_name}\" comment=\"{comment}\""
        
        filter_rules_data.append({
            "Rule Order": f"{rule_idx:02d}",
            "Chain": "forward",
            "Action": "accept",
            "Protocol": p,
            "Port": port,
            "Target Address-List": list_name,
            "Comment / Description": comment,
            "MikroTik Command (Copy & Paste)": command
        })
        rule_idx += 1

filter_rules_data.append({
    "Rule Order": "99",
    "Chain": "forward",
    "Action": "drop",
    "Protocol": "-",
    "Port": "-",
    "Target Address-List": "10.27.101.0/24 (Subnet)",
    "Comment / Description": "[99] DEFAULT DROP ALL TO .101 (Zero-Trust)",
    "MikroTik Command (Copy & Paste)": "add chain=forward action=drop dst-address=10.27.101.0/24 comment=\"[99] DEFAULT DROP ALL TO .101 (Zero-Trust)\""
})

# Create DataFrames
df_address = pd.DataFrame(address_list_data)
df_rules = pd.DataFrame(filter_rules_data)

# Export to Excel
excel_file = "MikroTik_Deployment_Plan.xlsx"
print(f"Exporting to {excel_file}...")

with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    df_address.to_excel(writer, sheet_name="1. Address Lists", index=False)
    df_rules.to_excel(writer, sheet_name="2. Filter Rules", index=False)

    # Styling
    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid") # Teal color for networking
    header_font = Font(color="FFFFFF", bold=True)
    command_font = Font(name="Consolas", color="0000FF")
    
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        # Headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # Format columns
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_str = str(cell.value)
                    if len(val_str) > max_length:
                        max_length = len(val_str)
                        
                    # Monospace font for command column
                    if "Command" in str(worksheet[f"{col_letter}1"].value):
                        if cell.row > 1:
                            cell.font = command_font
                except:
                    pass
            adjusted_width = min((max_length + 2), 100) # Max 100 width
            worksheet.column_dimensions[col_letter].width = adjusted_width

print("Done!")
