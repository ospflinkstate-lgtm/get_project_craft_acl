import json
import pandas as pd
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

STANDARD_PORTS = {
    3389, 8080, 8443, 3306, 5432, 27017, 6379, 9000, 1521, 1433,
    5060, 5061, 8000, 8001, 8002, 8081, 8888, 9200, 5900, 10050, 10051
}

def is_standard_port(port):
    if port is None: return False
    return port <= 1024 or port in STANDARD_PORTS

def get_service_name(port):
    m = {22:'SSH', 80:'HTTP', 443:'HTTPS', 3389:'RDP', 3306:'MySQL', 5432:'PostgreSQL', 53:'DNS', 161:'SNMP'}
    return m.get(port, "Unknown / Custom")

print("Loading data...")
try:
    with open('notlike101_traffic_combined.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
        hits = data.get('hits', [])
except Exception as e:
    print(f"Error loading traffic data: {e}")
    hits = []

try:
    with open('host_map.json', 'r', encoding='utf-8') as f:
        host_map = json.load(f)
except:
    host_map = {}

def get_hostname(ip):
    return host_map.get(ip, {}).get('name', 'Unknown Host')

print("Categorizing ports...")

# allowed_ports: dict of port -> { proto: set, target_ips: set, unique_sources: set }
allowed_ports = defaultdict(lambda: {'proto': set(), 'target_ips': set(), 'unique_sources': set()})

# dropped_traffic: dict of (dst_ip, dst_port) -> { proto: set, unique_sources: set }
dropped_traffic = defaultdict(lambda: {'proto': set(), 'unique_sources': set()})

for h in hits:
    dst_ip = h.get('dst_addr')
    src_ip = h.get('src_addr')
    dst_port = h.get('dst_port')
    proto = str(h.get('proto', 'TCP')).upper()

    if not dst_ip or not dst_ip.startswith('10.27.101.'): continue
    if src_ip == dst_ip: continue
    if dst_port is None: continue

    if is_standard_port(dst_port):
        allowed_ports[dst_port]['proto'].add(proto)
        allowed_ports[dst_port]['target_ips'].add(dst_ip)
        allowed_ports[dst_port]['unique_sources'].add(src_ip)
    else:
        key = (dst_ip, dst_port)
        dropped_traffic[key]['proto'].add(proto)
        dropped_traffic[key]['unique_sources'].add(src_ip)

print("Preparing DataFrames...")

# 1. Standard Ports Data
std_rows = []
for port in sorted(allowed_ports.keys()):
    stats = allowed_ports[port]
    std_rows.append({
        "Port Number": port,
        "Service Type": get_service_name(port),
        "Protocols": ", ".join(sorted(stats['proto'])),
        "Total Target Hosts (.101)": len(stats['target_ips']),
        "Total Unique Sources": len(stats['unique_sources']),
        "Reason": "Port <= 1024 or in Pre-approved Enterprise List"
    })
df_std = pd.DataFrame(std_rows)

# 2. Dropped Ports Data (Grouped by IP so they can review Custom Apps per Host)
drop_rows = []
for (dst_ip, dst_port) in sorted(dropped_traffic.keys(), key=lambda x: (int(x[0].split('.')[3]), x[1])):
    stats = dropped_traffic[(dst_ip, dst_port)]
    drop_rows.append({
        "Target Host IP": dst_ip,
        "Hostname": get_hostname(dst_ip),
        "Dropped Port Number": dst_port,
        "Protocols": ", ".join(sorted(stats['proto'])),
        "Unique Sources Count": len(stats['unique_sources']),
        "Potential Reason": "Ephemeral Port (Return Traffic) or Unknown Custom App",
        "Review Required": "Check if this is a Custom App"
    })
df_drop = pd.DataFrame(drop_rows)

# --- Export Excel (.xlsx) with Styling ---
excel_file = "Port_Categorization_Reference.xlsx"
print(f"Exporting to {excel_file}...")

with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    df_std.to_excel(writer, sheet_name="1. Standard Ports (Allowed)", index=False)
    df_drop.to_excel(writer, sheet_name="2. Dropped Ports (Review)", index=False)

    workbook = writer.book
    
    # Styles
    green_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    red_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Format Sheet 1 (Standard)
    ws1 = writer.sheets["1. Standard Ports (Allowed)"]
    for cell in ws1[1]:
        cell.fill = green_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ws1.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws1.column_dimensions[col_letter].width = min(max_length + 2, 60)

    # Format Sheet 2 (Dropped)
    ws2 = writer.sheets["2. Dropped Ports (Review)"]
    for cell in ws2[1]:
        cell.fill = red_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ws2.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws2.column_dimensions[col_letter].width = min(max_length + 2, 60)

print("Done!")
